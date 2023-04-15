# # import pandas as pd
# #
# # def excel(dict):
# #     df = pd.DataFrame(dict, index=[0, 1, 2, 3])
# #     # wr= pd.ExcelWriter('Output.xlsx', engine='xlsxwriter')
# #     reader = pd.read_excel('Output.xlsx', engine='openpyxl')
# #     writer = pd.ExcelWriter('Output.xlsx', engine='xlsxwriter', mode='a', if_sheet_exists="overlay")
# #
# #     df.to_excel(writer, index=False, sheet_name="Sheet1" ,header=False, startrow=len(reader)+1, engine='xlsxwriter')
# #
# #     writer.close()
#
# import pandas as pd
# from openpyxl.reader.excel import load_workbook
#
#
# def excel(dict):
#     df = pd.DataFrame(dict, index=[0, 1, 2, 3])
#     df.index = range(1, len(df) + 1)  # Set index explicitly
#     writer = pd.ExcelWriter('Output.xlsx', engine='openpyxl', mode='a')
#
#     writer.book = load_workbook('Output.xlsx')
#     writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
#     reader = pd.read_excel('Output.xlsx')
#     df.to_excel(writer, index=True, header=False, startrow=len(reader) + 1)
#
#     writer.save()
#     writer.close()

# import pandas as pd
#
# # Your dictionary
# my_dict = {'key1': 'value1', 'key2': 'value2', 'key3': 'value3'}
#
# # Create an empty dataframe to store the dictionary values
# df = pd.DataFrame(columns=['Key', 'Value'])
#
# # Iterate over the dictionary
# for key, value in my_dict.items():
#     # Append the key-value pair to the dataframe
#     df = df.append({'Key': key, 'Value': value}, ignore_index=True)
#
# # Write the dataframe to an Excel file
# df.to_excel('output.xlsx', index=False)
#


# import openpyxl
# wb = openpyxl.Workbook()
# ws = wb.active
# data = (
# ("Product","Cost Price","Selling Price"),
# ("earpod",90, 50),
# ("laptop", 3000, 8200),
# ("smartphone", 5100, 7200)
# )
# for i in data:
#    ws.append(i)
# wb.save('output.xlsx')

# ...Create a file
# import pandas as pd
#
# # create file
# writer = pd.ExcelWriter('iteration.xlsx', engine='xlsxwriter')
#
# # close file
# writer.close()
from testScripts import testVideo
from openpyxl import Workbook
book = Workbook()

sheet = book.active

rows = (testVideo.dict)

for row in rows:
    sheet.append(row)

book.save("output1.xlsx")